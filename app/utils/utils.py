from datetime import datetime, timedelta
import pandas as pd
from create_bot import (admins,
                        registrators,
                        doctors,
                        messeges,
                        day_names,
                        m_final)
import io
from openpyxl import Workbook
from aiogram.types import BufferedInputFile
from database.models import Appointments, Logs, Users


def get_day_by_index(index, short=False):
    if short:
        return day_names['short'][index]
    return day_names['full'][index]


def get_doctor_name_by_id(doctor_id: int, case='nom'):
    return doctors[case][doctor_id]


def to_datetime(date='01.01.2000', time='00:00'):
    return datetime.strptime(f'{date} {time}', '%d.%m.%Y %H:%M')


def handle_dates(dates: list[datetime]) -> dict:
    '''Функция превращает список дат в словарь, где ключем выступает
    месяц и год а значением выступает другой словарь, внутри которого
    лежат даты доступные в месяце, а так же следующий месяц,
    в котором есть даты, и предыдущий месяц, в котором есть даты'''
    if len(dates) == 0:
        return None
    dates = (pd.Series(dates)
             .astype('datetime64[ns]')
             .dt.normalize()
             .sort_values()
             .drop_duplicates()
             .rename('date')
             .to_frame()
             .assign(month=lambda x: x['date'].dt.month)
             .assign(year=lambda x: x['date'].dt.year)
             )
    unique_dates = (dates
                    [['month', 'year']]
                    .drop_duplicates()
                    .to_numpy()
                    )
    unique_dates = [(int(i), int(j)) for i, j in unique_dates]

    dates_dict = dict()
    for index, (month, year) in enumerate(unique_dates):
        dates_dict[(month, year)] = dict()
        if index == 0:
            dates_dict[(month, year)]['prev'] = None, None
        else:
            dates_dict[(month, year)]['prev'] = tuple(
                unique_dates[index - 1])
        if index == (len(unique_dates) - 1):
            dates_dict[(month, year)]['next'] = None, None
        else:
            dates_dict[(month, year)]['next'] = tuple(
                unique_dates[index + 1])
        dates_dict[(month, year)]['av_dates'] = (
            dates.query('year == @year')
            .query('month == @month')['date'].tolist())
    return dates_dict, unique_dates[0]


def get_date_inds(dates, include_sunday):
    date_range = pd.date_range(dates[0], dates[-1], freq='1 D')
    if not include_sunday:
        date_range = [date for date in date_range if date.weekday() != 6]
    ind = [date in dates for date in date_range]
    return date_range, ind


def get_curren_month_year():
    date = datetime.now()
    two_days_cond = ((date.hour >= 13)
                     and (date.minute >= 30)
                     and (date.weekday() == 6)
                     )

    one_day_cond = (((date.hour >= 17) and (date.minute >= 30))
                    or (date.weekday() == 6)
                    )

    if two_days_cond:
        date = pd.Timestamp(date.date() + timedelta(days=2))

    elif one_day_cond:
        date = pd.Timestamp(date.date() + timedelta(days=1))
    else:
        date = pd.Timestamp(date.date())
    return date.month, date.year


def get_dates_by_month_year(month, year, include_sunday=True):
    start_date = pd.Timestamp(datetime(day=1, month=month, year=year))
    end_date = start_date + pd.offsets.MonthEnd(0)
    date_range = pd.date_range(start_date, end_date)
    next_month = end_date + pd.offsets.MonthEnd(1)
    prev_month = end_date - pd.offsets.MonthEnd(1)

    if include_sunday:
        return (list(date_range),
                (next_month.month, next_month.year),
                (prev_month.month, prev_month.year))

    return ([i for i in date_range if i.weekday() != 6],
            (next_month.month, next_month.year),
            (prev_month.month, prev_month.year))


def get_other_doctor_id(doctor_id: int):
    od = {0: [], 1: [2], 2: [1], 3: [4], 4: [3], 5: []}
    return od[doctor_id]


def get_reply(q_type: str,
              doctor_id_from: int = None,
              doctor_id: int = None,
              doctor_true: int = None,
              doctor_true_from: int = None,
              date_from: str = None,
              date: str = None,
              time_from: str = None,
              time: str = None,
              time_start: str = None,
              time_end: str = None,
              regular_start_time: str = None,
              regular_end_time: str = None,
              weekend_start_time: str = None,
              weekend_end_time: str = None,
              treatment: bool = None,
              user_id: int = None,
              user_name: str = None,
              ve_days: list[int] = None,
              weekend_days: list[int] = None,
              days_without_work: list[int] = None,
              marks: list = None,
              visited: bool = None,
              banned: bool = None,
              short_weekday: bool = False,
              doctor_case: str = 'nom',
              ):
    '''Функция позволяет автоматически генерировать ответ по шаблону'''
    reply = messeges[q_type]

    if doctor_id_from is not None:
        reply += '\n\n<b>Текущая запись:</b>'
        if doctor_id_from == 5:
            reply += '\n<b>Лечение</b>'
        elif doctor_id_from == 0:
            doctor_name = get_doctor_name_by_id(
                doctor_id=doctor_id_from, case=doctor_case)
            reply += f'\nВыбранный врач: <b>{doctor_name}</b>'
            if ((doctor_true_from is not None)
               and (doctor_true_from != doctor_id_from)):
                doctor_name = get_doctor_name_by_id(
                        doctor_id=doctor_true_from, case=doctor_case)
                reply += f'\nВрач: <b>{doctor_name}</b>'
        else:
            doctor_name = get_doctor_name_by_id(
                doctor_id=doctor_id_from, case=doctor_case)
            reply += f'\nВрач: <b>{doctor_name}</b>'

    if date_from is not None:
        weekday = get_day_by_index(
            to_datetime(date_from).weekday(), short=short_weekday)
        reply += f'\nДата: <b>{date_from} ({weekday})</b>'

    if time_from is not None:
        reply += f'\nВремя: <b>{time_from}</b>'

    if user_id and (q_type in ['move_app', 'move_app_ext']):
        reply += f'\nid: <code>{user_id}</code>'
        if user_id in admins:
            reply += ' (владелец)'
        elif user_id in registrators:
            reply += ' (регистратор)'
        if user_name:
            reply += f', контакт: {user_name}'

    if doctor_id is not None:
        if q_type in ['move_user_app', 'move_app', 'move_app_ext']:
            reply += '\n\n<b>Новая запись</b>:'
        else:
            reply += '\n'

        if doctor_id == 5:
            reply += '\n<b>Лечение</b>'
            if q_type in ['make_user_app']:
                reply += '\n\n<i><b>Пожалуйста, не записывайтесь на лечение, '
                reply += 'если оно вам не назначено заранее</b></i>\n'
        elif doctor_id == 0:
            doctor_name = get_doctor_name_by_id(
                doctor_id=doctor_id, case=doctor_case)
            reply += f'\nВыбранный врач: <b>{doctor_name}</b>'
            if (doctor_true is not None) and (doctor_true != doctor_id):
                doctor_name = get_doctor_name_by_id(
                        doctor_id=doctor_true, case=doctor_case)
                reply += f'\nВрач: <b>{doctor_name}</b>'
        else:
            doctor_name = get_doctor_name_by_id(
                doctor_id=doctor_id, case=doctor_case)
            reply += f'\nВрач: <b>{doctor_name}</b>'

    if date is not None:
        if q_type in ['add_week', 'remove_week']:
            week_start = to_datetime(date)
            week_end = week_start + timedelta(days=6)
            week_start = week_start.strftime('%d.%m.%Y')
            week_end = week_end.strftime('%d.%m.%Y')
            reply += f'\n\nНеделя с <b>{week_start}</b> по <b>{week_end}</b>'
        else:
            weekday = get_day_by_index(
                to_datetime(date).weekday(), short=short_weekday)
            reply += f'\nДата: <b>{date} ({weekday})</b>'

    if time is not None:
        reply += f'\nВремя: <b>{time}</b>'

    if time_start is not None:
        reply += f'\nВремя начала приёма: <b>{time_start}</b>'
    if time_end is not None:
        reply += f'\nВремя окончания приёма: <b>{time_end}</b>'

    if regular_start_time and regular_end_time:
        reply += '\nВремя приёма в обычные дни: '
        reply += f'<b>{regular_start_time}</b> - <b>{regular_end_time}</b>'

    if weekend_start_time and weekend_end_time:
        reply += '\nВремя приёма в сокращённые дни: '
        reply += f'<b>{weekend_start_time}</b> - <b>{weekend_end_time}</b>'

    if treatment is not None:
        if treatment:
            reply += '\nДобавить запись на лечение: <b>да</b>'
        else:
            reply += '\nДобавить запись на лечение: <b>нет</b>'

    if ve_days is not None:
        ve_days = sorted(ve_days)
        if len(ve_days) == 1:
            reply += '\nДень когда принимает Виктор Евгеньевич:<b>'
            reply += f' {get_day_by_index(ve_days[0])}</b>'
        elif len(ve_days) > 1:
            reply += '\nДни когда принимает Виктор Евгеньевич:<b>'
            for day in ve_days:
                reply += ' ' + get_day_by_index(day) + ','
            reply = reply[:-1] + '</b>'
        else:
            reply += '\nУ Виктора Евгеньевича нет приёма на этой неделе'

    if weekend_days is not None:
        weekend_days = sorted(weekend_days)
        if len(weekend_days) == 1:
            reply += '\nСокращенный день:<b>'
            reply += f' {get_day_by_index(weekend_days[0])}</b>'
        elif len(weekend_days) > 1:
            reply += '\nСокращённые дни:<b>'
            for day in weekend_days:
                reply += ' ' + get_day_by_index(day) + ','
            reply = reply[:-1] + '</b>'
        else:
            reply += '\nНа этой неделе нет сокращённых дней'

    if days_without_work is not None:
        days_without_work = sorted(days_without_work)
        if len(days_without_work) == 1:
            reply += '\nВыходной день:<b>'
            reply += f' {get_day_by_index(days_without_work[0])}</b>'
        elif len(days_without_work) > 1:
            reply += '\nВыходные дни:<b>'
            for day in weekend_days:
                reply += ' ' + get_day_by_index(day) + ','
            reply = reply[:-1] + '</b>'
        else:
            reply += '\nНа этой неделе нет выходных дней'

    if user_id and (q_type not in ['move_user_app', 'move_app',
                                   'move_app_ext']):
        reply += f'\nid: <code>{user_id}</code>'
        if user_id in admins:
            reply += ' (владелец)'
        elif user_id in registrators:
            reply += ' (регистратор)'
        if user_name:
            reply += f', контакт: {user_name}'

    if visited is not None:
        if visited:
            reply += '\nПользователь приходил на приём'
        else:
            reply += '\n<b>Пользователь не приходил на приём</b>'

    if banned:
        reply += '\nПользователь <b>заблокирован</b>'

    if marks:
        reply += '\n<b>Пометки:</b>'
        for ind, note in enumerate(marks, 1):
            action_time = note.action_time.strftime('%d.%m.%Y')
            note = note.action.replace('\n', '').replace('  ', ' ').strip()
            reply += f'\n  {ind}. {action_time}, {note}'
    return reply + '\n'


def get_success_messege(user_id, user_name, status_code):
    text = f'<b>Успешно!</b>\n<b>{m_final[status_code]}</b>'
    if (status_code in [12, 19, 27]) and user_id:
        text += '\nНа данное время уже записан человек:'
    elif status_code == 14:
        text += '\nПользователь с которым совершено действие:'
    elif user_id:
        text += '\nНа данное время был записан человек:'
    if user_id:
        text += f'\nid: <code>{user_id}</code>'
    if user_id in admins:
        text += ' (владелец)'
    elif user_id in registrators:
        text += ' (регистратор)'
    if user_name:
        text += f', контакт: {user_name}'
    return text + '\n'


def get_warning_message(user_id, user_name, status_code):
    text = f'<b>Предупреждение!</b>\n{m_final[status_code]}'
    if user_id:
        text += '\nНа данное время сейчас записан человек'
        text += f'\nid: <b><code>{user_id}</code></b>'
    if user_id in admins:
        text += ' (владелец)'
    elif user_id in registrators:
        text += ' (регистратор)'
    if user_name:
        text += f', контакт: <b>{user_name}</b>'
    return text + '\n'


def reply_to_log(reply):
    log = []
    for string in reply.split('\n'):
        for word in ['лечение', 'врач', 'дата', 'время',
                     'текущая запись', 'новая запись']:
            if word in string.lower():
                log.append(string)

    return '\n'.join(log)


def generate_appointments_excel(apps: list[Appointments]) -> BufferedInputFile:
    wb = Workbook()
    ws = wb.active
    ws.title = "Appointments"

    headers = [
        "index",
        "Врач",
        "Время",
        "id",
        "Контакт",
        "Доступно (пользователь)",
        "Доступно (администратор)",
        "Пропустил"
    ]
    ws.append(headers)

    for app in apps:
        ws.append([
            app.id,
            get_doctor_name_by_id(app.doctor_id),
            app.time.strftime("%d.%m.%Y %H:%M") if app.time else None,
            app.user_id,
            app.user_name,
            "Да" if app.available_from_user else "Нет",
            "Да" if app.available_from_admin else "Нет",
            "Да" if app.skipped else "Нет"
        ])

    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return BufferedInputFile(buffer.getvalue(), filename="записи.xlsx")


def generate_logs_excel(logs: list[Logs]) -> BufferedInputFile:
    wb = Workbook()
    ws = wb.active
    ws.title = "Logs"

    headers = [
        "index",
        "id",
        "Тип действия",
        "Действие",
        "Время действия"
    ]
    ws.append(headers)

    for log in logs:
        ws.append([
            log.id,
            log.user_id,
            log.action_type,
            log.action,
            (log.action_time.strftime("%d.%m.%Y %H:%M")
             if log.action_time else None)
        ])

    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return BufferedInputFile(buffer.getvalue(), filename="действия.xlsx")


def generate_users_excel(users: list[Users]) -> BufferedInputFile:
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # Заголовки столбцов
    headers = [
        "id",
        "Контакт",
        "Заблокирован"
    ]
    ws.append(headers)

    for user in users:
        ws.append([
            user.user_id,
            user.user_name,
            "Да" if user.banned else "Нет"
        ])

    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return BufferedInputFile(buffer.getvalue(), filename="пользователи.xlsx")
